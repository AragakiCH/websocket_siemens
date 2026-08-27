# -*- coding: utf-8 -*-
"""
excel.py
========
Generación de ficheros Excel (.xlsx) con datos de PLC, usando **openpyxl**.

Una sola función pública, `construir_excel()`, que sirve para las dos fuentes:

  * una GRABACIÓN en vivo (muestreo a intervalo fijo), y
  * el HISTÓRICO ya guardado en base de datos.

Ambas llegan aquí como la misma lista de muestras, así que el fichero resultante
es idéntico venga de donde venga. El que lo abre no tiene que aprender dos
formatos.

Estructura del fichero:

  ┌───────────────┬──────────────────────────────────────────────────────────┐
  │ Información   │ Metadatos: PLCs, tags, rango, intervalo, nº de filas...   │
  │ Datos         │ PIVOTADO: una fila por instante, una columna por tag      │
  │ Estadísticas  │ Mín / máx / media / desviación / muestras por tag         │
  │ Tendencia     │ Gráfico de líneas de las variables numéricas             │
  └───────────────┴──────────────────────────────────────────────────────────┘

Por qué **pivotado** y no el formato estrecho de la base de datos: quien abre un
Excel casi siempre quiere graficar o hacer una tabla dinámica, y para eso
necesita una columna por variable. El formato estrecho (ts, tag, valor) es
excelente para almacenar y pésimo para analizar a mano.
"""
from __future__ import annotations

import io
import logging
import re
import statistics
from datetime import datetime, timezone
from typing import Any, Dict, Iterable, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger("excel")

# --- Paleta y estilos ------------------------------------------------- #
AZUL = "1F4E79"          # cabeceras
AZUL_CLARO = "DCE6F1"    # bandas
GRIS = "F2F2F2"
BLANCO = "FFFFFF"

FUENTE_CABECERA = Font(bold=True, color=BLANCO, size=11)
FUENTE_TITULO = Font(bold=True, size=14, color=AZUL)
FUENTE_ETIQUETA = Font(bold=True, size=10)
RELLENO_CABECERA = PatternFill("solid", fgColor=AZUL)
RELLENO_BANDA = PatternFill("solid", fgColor=AZUL_CLARO)
RELLENO_GRIS = PatternFill("solid", fgColor=GRIS)
CENTRADO = Alignment(horizontal="center", vertical="center", wrap_text=True)
IZQUIERDA = Alignment(horizontal="left", vertical="center")

_borde_fino = Side(style="thin", color="BFBFBF")
BORDE = Border(left=_borde_fino, right=_borde_fino,
               top=_borde_fino, bottom=_borde_fino)

FORMATO_FECHA = "yyyy-mm-dd hh:mm:ss"
FORMATO_NUM = "0.000"

# Límite de columnas del gráfico: más de 15 series es ilegible.
MAX_SERIES_GRAFICO = 15
# Excel admite ~1.048.576 filas; se corta antes para no generar ficheros
# imposibles de abrir.
MAX_FILAS = 500_000


# ====================================================================== #
# Utilidades
# ====================================================================== #
def _a_datetime(valor: Any) -> Optional[datetime]:
    """Convierte a datetime NAIVE (Excel no maneja zonas horarias)."""
    if valor is None:
        return None
    if isinstance(valor, datetime):
        dt = valor
    else:
        texto = str(valor).strip().replace("Z", "+00:00")
        try:
            dt = datetime.fromisoformat(texto)
        except ValueError:
            try:
                dt = datetime.strptime(texto[:19], "%Y-%m-%d %H:%M:%S")
            except ValueError:
                return None
    # Excel no entiende tzinfo: se pasa a UTC y se quita.
    if dt.tzinfo is not None:
        dt = dt.astimezone(timezone.utc).replace(tzinfo=None)
    return dt


def _nombre_hoja(nombre: str) -> str:
    """Sanea un nombre de hoja: Excel prohíbe : \\ / ? * [ ] y 31 caracteres."""
    limpio = re.sub(r"[:\\/?*\[\]]", "_", nombre or "Hoja")
    return limpio[:31] or "Hoja"


def _ajustar_columnas(hoja, anchos: Dict[int, int]) -> None:
    for idx, ancho in anchos.items():
        hoja.column_dimensions[get_column_letter(idx)].width = ancho


def _auto_ancho(hoja, max_col: int, minimo: int = 10, maximo: int = 30) -> None:
    """Ajusta el ancho a la longitud del contenido, acotado."""
    for col in range(1, max_col + 1):
        largo = minimo
        for celda in hoja[get_column_letter(col)][:200]:   # muestra de 200
            if celda.value is not None:
                largo = max(largo, len(str(celda.value)) + 2)
        hoja.column_dimensions[get_column_letter(col)].width = min(largo, maximo)


# ====================================================================== #
# Pivote
# ====================================================================== #
def pivotar(
    muestras: Iterable[dict],
    columna_ts: str = "ts",
    columna_tag: str = "tag",
    columna_valor: str = "valor",
    columna_plc: str = "plc",
    incluir_plc_en_columna: bool = True,
) -> Tuple[List[datetime], List[str], Dict[datetime, Dict[str, Any]]]:
    """
    Convierte muestras estrechas en una rejilla pivotada.

    Devuelve (instantes ordenados, nombres de columna, celdas[ts][columna]).

    Si dos PLCs tienen un tag con el mismo nombre, se prefija con el PLC para
    que no se pisen: es justo el caso de tener un Siemens y un Rexroth con la
    misma variable.
    """
    celdas: Dict[datetime, Dict[str, Any]] = {}
    columnas: Dict[str, None] = {}   # dict para conservar el orden de aparición
    nombres_por_tag: Dict[str, set] = {}

    filas: List[Tuple[datetime, str, str, Any]] = []
    for m in muestras:
        ts = _a_datetime(m.get(columna_ts))
        if ts is None:
            continue
        tag = str(m.get(columna_tag) or "")
        plc = str(m.get(columna_plc) or "")
        # El valor puede venir en 'valor', o separado en num/texto (BD).
        if columna_valor in m:
            valor = m.get(columna_valor)
        else:
            valor = m.get("valor_num")
            if valor is None:
                valor = m.get("valor_texto")
        filas.append((ts, plc, tag, valor))
        nombres_por_tag.setdefault(tag, set()).add(plc)

    # Solo se prefija el PLC cuando de verdad hace falta (o si se fuerza).
    for ts, plc, tag, valor in filas:
        if incluir_plc_en_columna and plc and len(nombres_por_tag.get(tag, ())) > 1:
            columna = f"{plc} | {tag}"
        else:
            columna = tag
        columnas.setdefault(columna, None)
        celdas.setdefault(ts, {})[columna] = valor

    instantes = sorted(celdas.keys())
    if len(instantes) > MAX_FILAS:
        instantes = instantes[-MAX_FILAS:]
        logger.warning("Export recortado a las %d filas más recientes.", MAX_FILAS)

    return instantes, list(columnas.keys()), celdas


# ====================================================================== #
# Hojas
# ====================================================================== #
def _hoja_info(wb: Workbook, info: Dict[str, Any]) -> None:
    """Hoja de metadatos: da trazabilidad al informe."""
    hoja = wb.create_sheet("Información", 0)
    hoja.sheet_view.showGridLines = False

    hoja["A1"] = info.get("titulo", "Exportación de datos de PLC")
    hoja["A1"].font = FUENTE_TITULO
    hoja.merge_cells("A1:B1")

    fila = 3
    for etiqueta, valor in info.get("campos", []):
        hoja.cell(fila, 1, etiqueta).font = FUENTE_ETIQUETA
        hoja.cell(fila, 1).fill = RELLENO_GRIS
        hoja.cell(fila, 1).border = BORDE
        celda = hoja.cell(fila, 2, valor)
        celda.border = BORDE
        celda.alignment = IZQUIERDA
        if isinstance(valor, datetime):
            celda.number_format = FORMATO_FECHA
        fila += 1

    # Lista de tags exportados, si viene.
    tags = info.get("tags") or []
    if tags:
        fila += 1
        hoja.cell(fila, 1, "Variables exportadas").font = FUENTE_CABECERA
        hoja.cell(fila, 1).fill = RELLENO_CABECERA
        hoja.cell(fila, 2, "").fill = RELLENO_CABECERA
        fila += 1
        for i, t in enumerate(tags, 1):
            hoja.cell(fila, 1, i).alignment = CENTRADO
            hoja.cell(fila, 1).border = BORDE
            hoja.cell(fila, 2, t).border = BORDE
            fila += 1

    _ajustar_columnas(hoja, {1: 28, 2: 55})


def _hoja_datos(
    wb: Workbook,
    instantes: List[datetime],
    columnas: List[str],
    celdas: Dict[datetime, Dict[str, Any]],
    nombre: str = "Datos",
) -> Tuple[Any, int, int]:
    """
    Hoja principal, PIVOTADA: fila = instante, columna = variable.

    Devuelve (hoja, nº de filas de datos, nº de columnas) para el gráfico.
    """
    hoja = wb.create_sheet(_nombre_hoja(nombre))

    # Cabecera
    hoja.cell(1, 1, "Fecha y hora")
    for i, col in enumerate(columnas, start=2):
        hoja.cell(1, i, col)
    for i in range(1, len(columnas) + 2):
        c = hoja.cell(1, i)
        c.font = FUENTE_CABECERA
        c.fill = RELLENO_CABECERA
        c.alignment = CENTRADO
        c.border = BORDE
    hoja.row_dimensions[1].height = 28

    # Datos
    for f, ts in enumerate(instantes, start=2):
        celda_ts = hoja.cell(f, 1, ts)
        celda_ts.number_format = FORMATO_FECHA
        celda_ts.border = BORDE
        fila_datos = celdas.get(ts, {})
        for i, col in enumerate(columnas, start=2):
            valor = fila_datos.get(col)
            celda = hoja.cell(f, i, valor)
            celda.border = BORDE
            if isinstance(valor, bool):
                # Los booleanos se escriben como texto legible, no TRUE/FALSE.
                celda.value = "ON" if valor else "OFF"
                celda.alignment = CENTRADO
            elif isinstance(valor, (int, float)):
                celda.number_format = FORMATO_NUM if isinstance(valor, float) else "0"
        # Banda alterna para leer mejor filas largas.
        if f % 2 == 0:
            for i in range(1, len(columnas) + 2):
                hoja.cell(f, i).fill = RELLENO_BANDA

    # Congelar la cabecera y la columna de tiempo + autofiltro.
    hoja.freeze_panes = "B2"
    if instantes:
        hoja.auto_filter.ref = (
            f"A1:{get_column_letter(len(columnas) + 1)}{len(instantes) + 1}"
        )
    _auto_ancho(hoja, len(columnas) + 1, minimo=12, maximo=26)
    hoja.column_dimensions["A"].width = 21

    return hoja, len(instantes), len(columnas)


def _hoja_estadisticas(
    wb: Workbook,
    columnas: List[str],
    celdas: Dict[datetime, Dict[str, Any]],
    instantes: List[datetime],
) -> None:
    """Mín/máx/media/desviación por variable. Solo tiene sentido en numéricas."""
    hoja = wb.create_sheet("Estadísticas")
    cabecera = ["Variable", "Muestras", "Mínimo", "Máximo", "Media",
                "Desv. típica", "Primer valor", "Último valor"]
    for i, texto in enumerate(cabecera, start=1):
        c = hoja.cell(1, i, texto)
        c.font = FUENTE_CABECERA
        c.fill = RELLENO_CABECERA
        c.alignment = CENTRADO
        c.border = BORDE
    hoja.row_dimensions[1].height = 26

    fila = 2
    for col in columnas:
        valores: List[float] = []
        primero = ultimo = None
        for ts in instantes:
            v = celdas.get(ts, {}).get(col)
            if v is None:
                continue
            if primero is None:
                primero = v
            ultimo = v
            if isinstance(v, bool):
                valores.append(1.0 if v else 0.0)
            elif isinstance(v, (int, float)):
                valores.append(float(v))

        hoja.cell(fila, 1, col).border = BORDE
        hoja.cell(fila, 2, len(valores) if valores else 0).border = BORDE
        if valores:
            datos = [
                min(valores), max(valores),
                statistics.fmean(valores),
                statistics.pstdev(valores) if len(valores) > 1 else 0.0,
            ]
            for i, v in enumerate(datos, start=3):
                c = hoja.cell(fila, i, round(v, 4))
                c.number_format = FORMATO_NUM
                c.border = BORDE
        else:
            # Variable de texto: no hay estadística numérica que dar.
            for i in range(3, 7):
                c = hoja.cell(fila, i, "—")
                c.alignment = CENTRADO
                c.border = BORDE

        for i, v in ((7, primero), (8, ultimo)):
            c = hoja.cell(fila, i, "ON" if v is True else "OFF" if v is False else v)
            c.border = BORDE
            if isinstance(v, float):
                c.number_format = FORMATO_NUM
        if fila % 2 == 0:
            for i in range(1, 9):
                hoja.cell(fila, i).fill = RELLENO_BANDA
        fila += 1

    hoja.freeze_panes = "A2"
    _auto_ancho(hoja, 8, minimo=12, maximo=32)


def _hoja_grafico(
    wb: Workbook,
    hoja_datos,
    num_filas: int,
    columnas: List[str],
    celdas: Dict[datetime, Dict[str, Any]],
    instantes: List[datetime],
) -> None:
    """Gráfico de líneas con las variables numéricas."""
    if num_filas < 2:
        return

    # Solo las columnas con algún número (una serie de texto rompe el gráfico).
    numericas: List[int] = []
    for i, col in enumerate(columnas, start=2):
        for ts in instantes[:200]:
            v = celdas.get(ts, {}).get(col)
            if isinstance(v, (int, float)) and not isinstance(v, bool):
                numericas.append(i)
                break
    if not numericas:
        return

    hoja = wb.create_sheet("Tendencia")
    hoja.sheet_view.showGridLines = False
    hoja["A1"] = "Tendencia de las variables numéricas"
    hoja["A1"].font = FUENTE_TITULO

    grafico = LineChart()
    grafico.title = "Evolución en el tiempo"
    grafico.style = 12
    grafico.y_axis.title = "Valor"
    grafico.x_axis.title = "Tiempo"
    grafico.height = 12
    grafico.width = 30

    recortadas = numericas[:MAX_SERIES_GRAFICO]
    for idx in recortadas:
        ref = Reference(hoja_datos, min_col=idx, min_row=1,
                        max_row=num_filas + 1)
        grafico.add_data(ref, titles_from_data=True)

    fechas = Reference(hoja_datos, min_col=1, min_row=2, max_row=num_filas + 1)
    grafico.set_categories(fechas)
    hoja.add_chart(grafico, "A3")

    if len(numericas) > MAX_SERIES_GRAFICO:
        hoja["A2"] = (f"Se muestran las primeras {MAX_SERIES_GRAFICO} variables "
                      f"numéricas de {len(numericas)}. El resto están en la "
                      f"hoja 'Datos'.")
        hoja["A2"].font = Font(italic=True, size=9)


# ====================================================================== #
# Función pública
# ====================================================================== #
def construir_excel(
    muestras: List[dict],
    titulo: str = "Exportación de datos de PLC",
    campos_info: Optional[List[Tuple[str, Any]]] = None,
    incluir_estadisticas: bool = True,
    incluir_grafico: bool = True,
    nombre_hoja_datos: str = "Datos",
) -> bytes:
    """
    Construye el .xlsx y lo devuelve como bytes, listo para descargar.

    `muestras` son filas estrechas con al menos `ts`, `tag` y el valor (en
    `valor`, o en `valor_num`/`valor_texto` si vienen de la base de datos).

    Se devuelve en memoria (no se escribe a disco) porque el fichero viaja
    directo al navegador; no hay que dejar basura en el servidor.
    """
    instantes, columnas, celdas = pivotar(muestras)

    wb = Workbook()
    wb.remove(wb.active)   # quitar la hoja vacía por defecto

    campos = list(campos_info or [])
    campos.extend([
        ("Filas exportadas", len(instantes)),
        ("Variables exportadas", len(columnas)),
        ("Primer registro", instantes[0] if instantes else "—"),
        ("Último registro", instantes[-1] if instantes else "—"),
        ("Generado", datetime.now().replace(microsecond=0)),
    ])
    _hoja_info(wb, {"titulo": titulo, "campos": campos, "tags": columnas})

    hoja_datos, num_filas, _ = _hoja_datos(
        wb, instantes, columnas, celdas, nombre_hoja_datos
    )

    if incluir_estadisticas and columnas:
        _hoja_estadisticas(wb, columnas, celdas, instantes)
    if incluir_grafico:
        _hoja_grafico(wb, hoja_datos, num_filas, columnas, celdas, instantes)

    # La hoja de datos es la que se ve al abrir el fichero.
    wb.active = wb.index(hoja_datos)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    logger.info("Excel generado: %d fila(s) x %d variable(s).",
                len(instantes), len(columnas))
    return buffer.getvalue()


def nombre_archivo(prefijo: str = "datos_plc") -> str:
    """Nombre de fichero con marca de tiempo, seguro para Windows."""
    sello = datetime.now().strftime("%Y%m%d_%H%M%S")
    limpio = re.sub(r"[^A-Za-z0-9_-]+", "_", prefijo).strip("_") or "datos_plc"
    return f"{limpio}_{sello}.xlsx"
