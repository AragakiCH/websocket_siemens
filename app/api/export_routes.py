# -*- coding: utf-8 -*-
"""
export_routes.py
================
Exportación a Excel (.xlsx) de los datos de los PLCs, desde dos fuentes:

  1. **En vivo** — se arranca una GRABACIÓN que muestrea los tags a intervalo
     fijo durante un periodo, y al terminar se descarga el fichero.
  2. **Desde la base de datos** — se exporta lo que ya guardó el historizador,
     filtrando por tag y rango de fechas.

Las dos producen el MISMO formato de Excel, así que quien lo abre no tiene que
aprender dos cosas:

  ┌───────────────┬──────────────────────────────────────────────────────────┐
  │ Información   │ Metadatos: origen, PLCs, tags, rango, intervalo, filas    │
  │ Datos         │ PIVOTADO: una fila por instante, una columna por variable │
  │ Estadísticas  │ Mín / máx / media / desviación / muestras por variable    │
  │ Tendencia     │ Gráfico de líneas de las variables numéricas             │
  └───────────────┴──────────────────────────────────────────────────────────┘

  GET    /export/tags                        -> tags grabables (para el selector).
  POST   /export/grabaciones                 -> arrancar una grabación.
  GET    /export/grabaciones                 -> listar con su estado.
  GET    /export/grabaciones/{id}            -> estado de una (para la barra de progreso).
  POST   /export/grabaciones/{id}/stop       -> parar antes de tiempo.
  DELETE /export/grabaciones/{id}            -> borrar y liberar memoria.
  GET    /export/grabaciones/{id}/excel      -> DESCARGAR el .xlsx.
  GET    /export/historico/excel             -> DESCARGAR desde la base de datos.
  POST   /export/consultas/{query_id}/excel  -> DESCARGAR el resultado de una consulta.
"""
from __future__ import annotations

import logging
from typing import List, Optional
from urllib.parse import quote

from fastapi import APIRouter, Body, Query, Request, Response
from pydantic import BaseModel, Field

from app.export.excel import construir_excel, nombre_archivo

logger = logging.getLogger("export_routes")

router = APIRouter()

TAG = ["Exportar a Excel"]

MIME_XLSX = ("application/vnd.openxmlformats-officedocument"
             ".spreadsheetml.sheet")


def _grab(request: Request):
    return request.app.state.grabador


def _respuesta_excel(datos: bytes, nombre: str) -> Response:
    """
    Devuelve el .xlsx como descarga.

    El nombre va en `filename*` codificado en UTF-8 para que no se rompa con
    acentos, y en `filename` sin acentos por compatibilidad con navegadores
    antiguos.
    """
    ascii_seguro = nombre.encode("ascii", "ignore").decode() or "export.xlsx"
    return Response(
        content=datos,
        media_type=MIME_XLSX,
        headers={
            "Content-Disposition": (
                f'attachment; filename="{ascii_seguro}"; '
                f"filename*=UTF-8''{quote(nombre)}"
            ),
            # Sin esto, el frontend no puede leer el nombre del fichero.
            "Access-Control-Expose-Headers": "Content-Disposition",
        },
    )


# ====================================================================== #
# Modelos
# ====================================================================== #
class NuevaGrabacion(BaseModel):
    """Cuerpo de POST /export/grabaciones."""

    grabacion_id: str = Field(
        ...,
        description="Identificador de la grabación. Ej: 'ensayo_arranque'.",
        examples=["ensayo_arranque"],
    )
    tags: List[str] = Field(
        default_factory=list,
        description="Tags a grabar, en formato `\"<plc_id>|<tag>\"`. "
                    "**Lista vacía = todos los tags disponibles.** "
                    "Los valores salen de `GET /export/tags`.",
        examples=[["192.168.50.1|DB_snap7.temperatura",
                   "192.168.100.31|PLC_PRG.AI_Sensor_mA"]],
    )
    intervalo_ms: int = Field(
        default=1000, ge=100, le=3_600_000,
        description="Cada cuánto se toma una muestra de TODOS los tags a la "
                    "vez. Mínimo 100 ms (por debajo, el PLC no publica más "
                    "rápido y solo saldrían filas repetidas).",
    )
    duracion_s: int = Field(
        default=60, ge=0, le=86_400,
        description="Cuánto dura la grabación en segundos. **0 = indefinida**, "
                    "hasta que se pare con `/stop`.",
    )
    nombre: str = Field(
        default="", description="Etiqueta legible; sale en la hoja Información.")


# ====================================================================== #
# Tags disponibles
# ====================================================================== #
@router.get(
    "/export/tags",
    tags=TAG,
    summary="Tags disponibles para grabar",
    description="Lista los tags de todos los PLCs conectados con su valor "
                "actual y la **clave compuesta** `\"plc|tag\"` ya montada, que "
                "es la que espera `POST /export/grabaciones`.\n\n"
                "Es la información de `GET /tags` preparada para el selector de "
                "variables de la vista.",
    responses={200: {"content": {"application/json": {"example": {
        "num_tags": 2,
        "tags": [
            {"clave": "192.168.50.1|DB_snap7.temperatura",
             "plc": "192.168.50.1", "tag": "DB_snap7.temperatura",
             "tipo": "Float", "valor_actual": 53.09},
            {"clave": "192.168.100.31|PLC_PRG.AI_Sensor_mA",
             "plc": "192.168.100.31", "tag": "PLC_PRG.AI_Sensor_mA",
             "tipo": "REAL", "valor_actual": 8.09},
        ],
    }}}}},
)
async def tags_grabables(request: Request) -> dict:
    tags = _grab(request).tags_disponibles()
    return {"num_tags": len(tags), "tags": tags}


# ====================================================================== #
# Grabaciones en vivo
# ====================================================================== #
@router.post(
    "/export/grabaciones",
    tags=TAG,
    summary="Iniciar una grabación en vivo",
    description="Empieza a muestrear los tags elegidos cada `intervalo_ms` "
                "durante `duracion_s` segundos. Al terminar (o al pararla a "
                "mano) se descarga el Excel con "
                "`GET /export/grabaciones/{id}/excel`.\n\n"
                "**Por qué muestreo a intervalo fijo y no cada cambio**: si se "
                "guardara cada cambio, cada variable tendría marcas de tiempo "
                "distintas y el Excel saldría lleno de huecos. Muestreando "
                "todas a la vez, comparten fila y el gráfico sale limpio.\n\n"
                "No abre ninguna conexión extra al PLC: usa el mismo flujo que "
                "alimenta el WebSocket.\n\n"
                "⚠️ Las muestras se acumulan en memoria (máx. 200.000). Para "
                "histórico permanente, usa el **Historizador**, que escribe en "
                "base de datos.",
    responses={200: {"content": {"application/json": {"examples": {
        "ok": {"summary": "Grabación en curso", "value": {
            "ok": True, "grabacion_id": "ensayo_arranque",
            "intervalo_ms": 1000, "duracion_s": 60, "num_tags": 2,
            "mensaje": "Grabación 'ensayo_arranque' en curso.",
        }},
        "aviso": {"summary": "Con tags que no existen", "value": {
            "ok": True, "grabacion_id": "ensayo_arranque",
            "intervalo_ms": 1000, "duracion_s": 60, "num_tags": 2,
            "tags_desconocidos": ["192.168.50.1|DB.noexiste"],
            "mensaje": "Grabación 'ensayo_arranque' en curso. Aviso: 1 tag(s) "
                       "no existen ahora mismo y saldrán vacíos.",
        }},
        "duplicada": {"summary": "Ya estaba grabando", "value": {
            "ok": False, "grabacion_id": "ensayo_arranque",
            "mensaje": "La grabación 'ensayo_arranque' ya está en curso. "
                       "Párala antes de volver a empezar.",
        }},
    }}}}},
)
async def iniciar_grabacion(
    request: Request,
    cuerpo: NuevaGrabacion = Body(..., openapi_examples={
        "seleccion": {
            "summary": "Variables elegidas, 1 minuto a 1 s",
            "description": "El caso normal: el usuario marca las variables en "
                           "la vista y graba un rato para analizar.",
            "value": {
                "grabacion_id": "ensayo_arranque",
                "nombre": "Ensayo de arranque",
                "tags": [
                    "192.168.50.1|DB_snap7.temperatura",
                    "192.168.50.1|DB_snap7.presion",
                    "192.168.100.31|PLC_PRG.AI_Sensor_mA",
                ],
                "intervalo_ms": 1000,
                "duracion_s": 60,
            },
        },
        "rapida": {
            "summary": "Alta resolución, 10 s a 100 ms",
            "description": "Para capturar un transitorio rápido. 100 ms es el "
                           "mínimo real del servidor OPC UA del S7-1500.",
            "value": {
                "grabacion_id": "transitorio",
                "nombre": "Transitorio de arranque",
                "tags": ["192.168.50.1|DB_snap7.temperatura"],
                "intervalo_ms": 100,
                "duracion_s": 10,
            },
        },
        "todos_indefinida": {
            "summary": "Todos los tags, hasta pararla a mano",
            "description": "Lista vacía = todas las variables disponibles. "
                           "`duracion_s: 0` = indefinida: hay que llamar a "
                           "`/stop` para terminarla.",
            "value": {
                "grabacion_id": "turno_manana",
                "nombre": "Turno de mañana",
                "tags": [],
                "intervalo_ms": 5000,
                "duracion_s": 0,
            },
        },
    }),
) -> dict:
    return _grab(request).iniciar(
        grabacion_id=cuerpo.grabacion_id, tags=cuerpo.tags,
        intervalo_ms=cuerpo.intervalo_ms, duracion_s=cuerpo.duracion_s,
        nombre=cuerpo.nombre,
    )


@router.get(
    "/export/grabaciones",
    tags=TAG,
    summary="Listar grabaciones",
    description="Todas las grabaciones con su estado. `estado` puede ser "
                "`grabando`, `terminada` o `detenida`.",
    responses={200: {"content": {"application/json": {"example": {
        "num_grabaciones": 1, "en_curso": 1, "tags_en_cache": 5,
        "grabaciones": [{
            "grabacion_id": "ensayo_arranque", "nombre": "Ensayo de arranque",
            "estado": "grabando", "tags": ["192.168.50.1|DB_snap7.temperatura"],
            "todos_los_tags": False, "num_tags": 1,
            "intervalo_ms": 1000, "duracion_s": 60,
            "inicio": "2026-07-30T18:00:00+00:00", "fin": None,
            "segundos_transcurridos": 23.4, "segundos_restantes": 36.6,
            "num_muestras": 23, "motivo_fin": "", "descargable": True,
        }],
    }}}}},
)
async def listar_grabaciones(request: Request) -> dict:
    return _grab(request).listar()


@router.get(
    "/export/grabaciones/{grabacion_id}",
    tags=TAG,
    summary="Estado de una grabación",
    description="Mismo contenido que el listado, para una sola. Sirve para "
                "pintar una barra de progreso: `segundos_restantes` y "
                "`num_muestras` se actualizan en cada llamada.",
)
async def estado_grabacion(request: Request, grabacion_id: str) -> dict:
    grabacion = _grab(request).obtener(grabacion_id)
    if grabacion is None:
        return {"ok": False, "mensaje": f"No existe la grabación '{grabacion_id}'."}
    salida = {"ok": True}
    salida.update(grabacion.estado_dict())
    return salida


@router.post(
    "/export/grabaciones/{grabacion_id}/stop",
    tags=TAG,
    summary="Parar una grabación",
    description="Termina antes de tiempo. Los datos capturados hasta ese "
                "momento **siguen disponibles** para descargar.",
    responses={200: {"content": {"application/json": {"example": {
        "ok": True, "grabacion_id": "ensayo_arranque", "estado": "terminada",
        "num_muestras": 142, "segundos": 47.3,
        "mensaje": "Grabación 'ensayo_arranque' detenida con 142 muestra(s). "
                   "Ya se puede descargar el Excel.",
    }}}}},
)
async def parar_grabacion(request: Request, grabacion_id: str) -> dict:
    return _grab(request).parar(grabacion_id)


@router.delete(
    "/export/grabaciones/{grabacion_id}",
    tags=TAG,
    summary="Borrar una grabación",
    description="Elimina la grabación y **libera su memoria**. Conviene "
                "hacerlo tras descargar el Excel, sobre todo con grabaciones "
                "largas.",
)
async def borrar_grabacion(request: Request, grabacion_id: str) -> dict:
    return _grab(request).borrar(grabacion_id)


@router.get(
    "/export/grabaciones/{grabacion_id}/excel",
    tags=TAG,
    summary="⬇️ Descargar el Excel de una grabación",
    description="Devuelve el fichero `.xlsx` como descarga directa.\n\n"
                "Se puede llamar con la grabación aún en curso: exporta lo "
                "capturado hasta ese momento.\n\n"
                "El nombre del fichero viaja en la cabecera "
                "`Content-Disposition` (ya expuesta por CORS), así que el "
                "frontend puede leerlo:\n\n"
                "```js\n"
                "const r = await fetch(`/export/grabaciones/${id}/excel`);\n"
                "const blob = await r.blob();\n"
                "const url = URL.createObjectURL(blob);\n"
                "const a = Object.assign(document.createElement('a'),\n"
                "  { href: url, download: 'grabacion.xlsx' });\n"
                "a.click(); URL.revokeObjectURL(url);\n"
                "```",
    response_class=Response,
    responses={
        200: {"content": {MIME_XLSX: {}},
              "description": "Fichero Excel con las hojas Información, Datos, "
                             "Estadísticas y Tendencia."},
        404: {"description": "La grabación no existe o no tiene muestras."},
    },
)
async def descargar_grabacion(request: Request, grabacion_id: str):
    grabacion = _grab(request).obtener(grabacion_id)
    if grabacion is None:
        return Response(
            content=f'{{"ok":false,"mensaje":"No existe la grabación '
                    f'\'{grabacion_id}\'."}}',
            media_type="application/json", status_code=404)
    if not grabacion.muestras:
        return Response(
            content='{"ok":false,"mensaje":"La grabación todavía no tiene '
                    'muestras."}',
            media_type="application/json", status_code=404)

    datos = construir_excel(
        grabacion.muestras,
        titulo=f"Grabación en vivo: {grabacion.nombre}",
        campos_info=[
            ("Origen", "Grabación en vivo (muestreo del PLC)"),
            ("Grabación", grabacion.grabacion_id),
            ("Estado", grabacion.estado),
            ("Intervalo de muestreo", f"{grabacion.intervalo_ms} ms"),
            ("Duración configurada",
             f"{grabacion.duracion_s} s" if grabacion.duracion_s else "Indefinida"),
            ("Duración real", f"{grabacion.segundos_transcurridos:.1f} s"),
            ("Muestras capturadas", len(grabacion.muestras)),
        ],
    )
    return _respuesta_excel(datos, nombre_archivo(grabacion.grabacion_id))


# ====================================================================== #
# Exportación desde la base de datos
# ====================================================================== #
@router.get(
    "/export/historico/excel",
    tags=TAG,
    summary="⬇️ Descargar el histórico de la BD en Excel",
    description="Exporta lo que ya guardó el **historizador** en base de "
                "datos, filtrando por tag y rango de fechas.\n\n"
                "A diferencia de la grabación (que vive en memoria y es "
                "temporal), aquí se puede exportar cualquier periodo pasado, "
                "incluso de hace meses.\n\n"
                "El fichero sale en el mismo formato pivotado: una fila por "
                "instante, una columna por variable.",
    response_class=Response,
    responses={
        200: {"content": {MIME_XLSX: {}}, "description": "Fichero Excel."},
        404: {"description": "El grupo no existe o no hay datos en ese rango."},
    },
)
async def descargar_historico(
    request: Request,
    grupo_id: str = Query(
        ...,
        description="Grupo de historización (de `GET /historian`).",
        examples=["proceso"],
    ),
    tag: Optional[str] = Query(
        default=None,
        description="Filtrar por un tag concreto, sin el prefijo del PLC. "
                    "Ej: `DB_snap7.temperatura`. Vacío = todos los del grupo.",
    ),
    desde: Optional[str] = Query(
        default=None,
        description="Fecha/hora inicial en ISO 8601. Ej: `2026-07-30T00:00:00`.",
    ),
    hasta: Optional[str] = Query(
        default=None, description="Fecha/hora final en ISO 8601."),
    limite: int = Query(
        default=10000, ge=1, le=100_000,
        description="Máximo de registros a leer de la base de datos.",
    ),
):
    historizador = request.app.state.historizador
    resultado = await historizador.leer(grupo_id, tag, desde, hasta, limite)

    if not resultado.get("ok"):
        import json
        return Response(content=json.dumps(resultado, ensure_ascii=False),
                        media_type="application/json", status_code=404)

    filas = resultado.get("filas") or []
    if not filas:
        return Response(
            content='{"ok":false,"mensaje":"No hay datos historizados en ese '
                    'rango."}',
            media_type="application/json", status_code=404)

    datos = construir_excel(
        filas,
        titulo=f"Histórico: {grupo_id}",
        campos_info=[
            ("Origen", "Base de datos (historizador)"),
            ("Grupo", grupo_id),
            ("Tabla", resultado.get("tabla", "")),
            ("Filtro de tag", tag or "Todos"),
            ("Desde", desde or "Sin límite"),
            ("Hasta", hasta or "Sin límite"),
            ("Registros leídos", resultado.get("num_filas", len(filas))),
            ("Recortado por el límite", "Sí" if resultado.get("truncado") else "No"),
        ],
    )
    return _respuesta_excel(datos, nombre_archivo(f"historico_{grupo_id}"))


@router.post(
    "/export/consultas/{query_id}/excel",
    tags=TAG,
    summary="⬇️ Descargar el resultado de una consulta en Excel",
    description="Ejecuta una consulta guardada (`POST /db/queries`) y devuelve "
                "el resultado como Excel. Sirve para dar un botón "
                "\"Exportar\" a cualquier widget de datos.\n\n"
                "Si el resultado tiene columnas `ts`/`tag`/`valor` se pivota "
                "igual que el resto; si es una tabla cualquiera (ej. producción "
                "por máquina), se vuelca tal cual.",
    response_class=Response,
    responses={
        200: {"content": {MIME_XLSX: {}}, "description": "Fichero Excel."},
        404: {"description": "La consulta no existe o no devolvió filas."},
    },
)
async def descargar_consulta(
    request: Request,
    query_id: str,
    parametros: dict = Body(
        default_factory=dict,
        description="Valores de los parámetros de la consulta.",
        examples=[{"desde": "2026-07-01"}],
    ),
):
    import json

    db = request.app.state.db_manager
    resultado = await db.ejecutar(query_id, parametros or {})
    if not resultado.get("ok"):
        return Response(content=json.dumps(resultado, ensure_ascii=False),
                        media_type="application/json", status_code=404)

    filas = resultado.get("filas") or []
    if not filas:
        return Response(
            content='{"ok":false,"mensaje":"La consulta no devolvió filas."}',
            media_type="application/json", status_code=404)

    columnas = resultado.get("columnas") or []
    campos = [
        ("Origen", "Consulta guardada"),
        ("Consulta", query_id),
        ("Conexión", resultado.get("db_id", "")),
        ("Parámetros", json.dumps(resultado.get("parametros") or {},
                                  ensure_ascii=False)),
        ("Filas devueltas", resultado.get("num_filas", len(filas))),
    ]

    # Si la consulta tiene forma de serie temporal, se pivota; si no, se
    # vuelca tal cual (una tabla normal no se debe pivotar).
    if "ts" in columnas and "tag" in columnas:
        datos = construir_excel(filas, titulo=f"Consulta: {query_id}",
                                campos_info=campos)
    else:
        datos = _excel_tabla_plana(filas, columnas, query_id, campos)

    return _respuesta_excel(datos, nombre_archivo(f"consulta_{query_id}"))


def _excel_tabla_plana(filas: list, columnas: list, titulo: str,
                       campos: list) -> bytes:
    """
    Vuelca una tabla cualquiera (sin pivotar) con el mismo estilo.

    Se usa cuando la consulta no es una serie temporal: por ejemplo "piezas por
    máquina", donde pivotar no tendría sentido.
    """
    import io
    from datetime import datetime

    from openpyxl import Workbook

    from app.export.excel import (BORDE, CENTRADO, FUENTE_CABECERA,
                                  RELLENO_BANDA, RELLENO_CABECERA, _auto_ancho,
                                  _hoja_info)

    wb = Workbook()
    wb.remove(wb.active)

    todos = list(campos) + [
        ("Generado", datetime.now().replace(microsecond=0)),
    ]
    _hoja_info(wb, {"titulo": f"Consulta: {titulo}", "campos": todos,
                    "tags": columnas})

    hoja = wb.create_sheet("Datos")
    for i, col in enumerate(columnas, start=1):
        c = hoja.cell(1, i, col)
        c.font = FUENTE_CABECERA
        c.fill = RELLENO_CABECERA
        c.alignment = CENTRADO
        c.border = BORDE
    hoja.row_dimensions[1].height = 26

    for f, fila in enumerate(filas, start=2):
        for i, col in enumerate(columnas, start=1):
            valor = fila.get(col)
            celda = hoja.cell(f, i, valor)
            celda.border = BORDE
            if isinstance(valor, bool):
                celda.value = "SÍ" if valor else "NO"
                celda.alignment = CENTRADO
        if f % 2 == 0:
            for i in range(1, len(columnas) + 1):
                hoja.cell(f, i).fill = RELLENO_BANDA

    hoja.freeze_panes = "A2"
    if filas:
        from openpyxl.utils import get_column_letter
        hoja.auto_filter.ref = (
            f"A1:{get_column_letter(len(columnas))}{len(filas) + 1}")
    _auto_ancho(hoja, len(columnas), minimo=12, maximo=40)
    wb.active = wb.index(hoja)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
